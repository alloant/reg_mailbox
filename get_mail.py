#!/bin/python

from getpass import getpass
from datetime import datetime
from tempfile import NamedTemporaryFile
import time
import re
from pathlib import Path
import ast

import pickle

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from synology_drive_api.drive import SynologyDrive

from syno_tools import txt2dict, copy_to, move_to, convert_to, EXT

TO_ARCHIVE = "Outbox Despacho/to_archive"

def get_ctr(folder):
    if folder[:7] == 'Mailbox':
        return folder[8:]
    elif folder[:10] == '8.Mailbox-':
        return folder[10:]


#ext = {'xls':'osheet','xlsx':'osheet','docx':'odoc'}

def folder_in_teams(folder,teams):
    fds = folder.split("/")[1:]
    for team in teams:
        tms = team['folder'].split("/")[1:]
        same = True
        key = team['type']
        for i,fd in enumerate(fds):
            if '@' in tms[i]:
                pt = tms[i].replace('@','')
                if fds[i][:len(pt)] == pt:
                    key = fds[i][len(pt):]
                else:
                    same = False
                    break
            else:
                if fds[i] != tms[i]:
                    same = False
                    break

        if same:
            nt = team.copy()
            nt['folder'] = team['folder'].replace('@',key)
            return same,key,nt

    return False,'asd',''


def get_notes_in_folders(PASS):
    config = txt2dict("config.txt")
    year = datetime.today().strftime('%Y')

    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd: 
        team_folders = synd.get_teamfolder_info()
        team_config = [ast.literal_eval(tm) for tm in config['teams'].split('|')]

        reg_notes = {}
        for folder in team_folders:
            mail_folder,key,team = folder_in_teams(f"/team-folders/{folder}",team_config)
            
            if mail_folder:# and key == 'gul':
                print(f"Checking folder {team['folder']}")
                try:
                    notes = synd.list_folder(team['folder'])['data']['items']
                    for note in notes:
                        print(f"    ->Found note {note['name']}")
                        reg_notes[note['name']] = team.copy()|{'source':key,'converted':False,'original':''}
                except:
                    print(f'Cannot get notes from {team["folder"]}')
                    continue

        return reg_notes

def create_register(ws,reg_notes):
    year = datetime.today().strftime('%Y')
    ws.append(['type','source','No','Year','Ref','Date','Content','Dept','Name','Original'])
    
    for name,note in dict(sorted(reg_notes.items())).items():
        num = re.findall('\d+',name.replace(note['source'],''))
        num = num[0] if num else ''
        
        #num = f"000{num[0]}"[-4:] if num else ''
        #if num and note['type'] == 'ctr in': num = num[1:]
        
        note['num'] = num

        note['year'] = year

        if note['type'] == 'r in':
            src = re.findall('\D+',name)
            note['source'] = src[0] if src else ''
        
        if note['converted']: continue
        
        nm = note['link'] if 'link' in note else name

        ws.append([note['type'],note['source'],num,year,'','','','',nm,note['original']])    
        ws[ws.max_row][5].value = datetime.today()
        ws[ws.max_row][5].number_format = 'dd/mm/yyyy;@'

        column_widths = [10,10,10,10,12,12,50,12,20,20]
        for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
            ws.column_dimensions[get_column_letter(i)].width = column_width

        for row in ws[1:ws.max_row]:  # skip the header
            for i,col in enumerate(column_widths):
                cell = row[i]             # column H
                cell.alignment = Alignment(horizontal='center')

def change_names(PASS,notes):
    config = txt2dict("config.txt")

    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        new_names = []
        for name,note in notes.items():
            try:
                new_name = name if note['source'] in name else f"{note['source']}_{name}"
                new_name = new_name.strip()
                #new_name = new_name.replace(' ','_')
                new_names.append([new_name,name])
                synd.rename_path(new_name,f"{note['folder']}/{name}")
                
            except Exception as err:
                print(f"ERROR Cannot change name of {name}: {err}")

        for new in new_names:
            notes[new[0]] = notes.pop(new[1])

def move_to_despacho(PASS,notes):
    config = txt2dict("config.txt")

    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        for name,note in notes.items():
            try:
                name_link = name
                f_id,p_link = move_to(synd,f"{note['folder']}/{name}",f"{config['despacho']}/{TO_ARCHIVE}")
                note['folder'] = f"{config['despacho']}/{TO_ARCHIVE}"
            except:
                print(f"Cannot move {name}")


def convert_files(PASS,notes):
    config = txt2dict("config.txt")

    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        new_names = []
        for name,note in notes.items():
            try:
                name_link = name
                note['converted'] = False
                # Here I check if I can convert the file to synology with the extension
                ext = Path(name).suffix[1:]
                if ext in EXT:
                    convert_to(synd,f"{config['despacho']}/{TO_ARCHIVE}/{name}")
                    #note['converted'] = True
                    note['original'] = name
                    new_names.append([f"{name[:-len(ext)]}{EXT[ext]}",name])
            except:
                print(f"Cannot convert {name}")

        for new in new_names:
            notes[new[0]] = notes.pop(new[1])



def copy_to_despacho(PASS,notes):
    config = txt2dict("config.txt")
    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        for name,note in notes.items():
            try:
                ext = Path(name).suffix[1:]
                if note['converted']:
                    name_link = f"{name[:-len(ext)]}{EXT[ext]}"
                else:
                    name_link = name
                
                f_id,p_link = copy_to(synd,f"{note['folder']}/{name_link}",f"{config['despacho']}/Inbox Despacho")

                chain = 'd/f'
                if ext in EXT.values():
                    chain = 'oo/r'

                link = f'=HYPERLINK("#dlink=/{chain}/{p_link}", "{name_link}")' if p_link != '' else name_link
                note['link'] = link
            except:
                print(f"Cannot copy {name}")

def upload_register(PASS,wb):
    config = txt2dict("config.txt")
    date = datetime.today().strftime('%Y-%m-%d')
    hour = datetime.today().strftime('%HH-%Mm')
    
    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        file = NamedTemporaryFile()
        wb.save(file)
        file.seek(0)
        file.name = f"despacho-{date}-{hour}.xlsx"
    
        print("Creating register file")
        uploaded = True
        
        try:
            ret_upload = synd.upload_file(file, dest_folder_path=f"{config['despacho']}/Inbox Despacho")
        except:
            print("Cannot upload register")
            wb.save(f"despacho-{date}-{hour}.xlsx")
            uploaded = False

        if uploaded:
            try:
                ret_convert = synd.convert_to_online_office(ret_upload['data']['display_path'],
                    delete_original_file=False,
                    conflict_action='autorename')
            except:
                print("Cannot convert file to Synology Office")

            return True
    
    print("Cannot upload register")
    wb.save(f"despacho-{date}-{hour}.xlsx")




def main():
    PASS = getpass()
    
    reg_notes = get_notes_in_folders(PASS)
    #input("go") 
    if reg_notes != {}:
        wb = Workbook()
        ws = wb.active
        
        try:
            change_names(PASS,reg_notes)
            move_to_despacho(PASS,reg_notes)
            convert_files(PASS,reg_notes)
            copy_to_despacho(PASS,reg_notes)
            create_register(ws,reg_notes)
            upload_register(PASS,wb)
        except:
            create_register(ws,reg_notes)
            upload_register(PASS,wb)
    
    input("Pulse Enter to continue")



if __name__ == '__main__':
    main()
