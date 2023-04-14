#!/bin/python

from getpass import getpass
from datetime import datetime
from tempfile import NamedTemporaryFile
import time
import re
from pathlib import Path
import ast


from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from synology_drive_api.drive import SynologyDrive

from syno_tools import txt2dict, copy_to, move_to, convert_to, EXT

TO_ARCHIVE = "Outbox Despacho/to_archive"

TITLES = ['type','source','No','Year','Ref','Date','Content','Dept','Name','Original'] 

def read_register(PASS):
    config = txt2dict("config.txt")
    
    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        register = synd.list_folder(f"{config['despacho']}/Outbox Despacho")['data']['items']
        pattern = re.compile("despacho-(\d\d-\d\d-\d\d\d\d)-(\d\dH-\d\dm).osheet")
        notes = {}
        
        for reg in register:
            #print(reg["name"],re.match(pattern,reg["name"]))
            #if re.match(pattern,reg["name"]):
            if reg["name"][:9] == "despacho-" and 'osheet' in reg['name']:
                try:
                    reg_file = synd.download_file(reg['display_path'])
                    wb = load_workbook(reg_file)
                    
                    for ws_name in wb.sheetnames:
                        ws = wb[ws_name]
                        
                        for row in ws.iter_rows(values_only=True):
                            if row[2] != '' and row[1] != '' and row[0] != 'type':
                                n = f"000{row[2]}"[-4:]
                                notes[f"{row[1]}_{n}"] = {'notes':[]}
                       
                        for row in ws.iter_rows(values_only=True):
                            if row[2] != '' and row[1] != '' and row[0] != 'type':
                                temp = {}
                                for i,title in enumerate(TITLES):
                                    if not title in temp: temp[title] = ''

                                    if not row[i] in ['',None]:
                                        temp[title] = row[i]

                                    if title == 'No':
                                        temp[title] = f"000{row[i]}"[-4:]
                                    if isinstance(row[i],datetime):
                                        temp[title] = row[i].strftime('%d/%m/%Y')

                                temp['main'] = True if temp['Content'] != '' else False

                                notes[f"{temp['source']}_{temp['No']}"]['notes'].append(temp.copy())
                except:
                    print(f"Cannot download {note['name']}")
                    continue

        return notes


def create_folder(synd,data,dest,num):
    if data['create_folder']:
        try:
            new_folder = synd.get_file_or_folder_info(f"{dest}/{num}")
            exist = True
        except:
            exist = False

        try:
            if not exist:
                rst = synd.create_folder(num, dest)
                p_link = rst['data']['permanent_link']
            else:
                p_link = new_folder['data']['permanent_link']

            dest = f"{dest}/{num}"

            return dest,p_link
        except:
            print(f"Cannot create folder {num}")
            return '',''
    else:
        return dest,''


def rename_note(synd,note_name,path,num):
        #name = note_name[:-2].split('","')[1]
        name = note_name
        try:
            ext = Path(name).suffix[1:]
            synd.rename_path(f"{num}.{ext}",f"{path}/{name}")
            name = f"{num}.{ext}"
        except:
            print("Cannot change the name")
        
        return name

        
def archive_notes(PASS,reg_notes):
    config = txt2dict("config.txt")
    path = f"{config['despacho']}/{TO_ARCHIVE}"
    
    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        for num,data in reg_notes.items():
            if data['Dept'] in [None,'']:
                data['link'] = ''
                continue

            dest = f"{config['archive']}/{data['type']} in {data['Year']}"
            dest,path_link = create_folder(synd,data,dest,num)
            
            for note in data['notes']:
                if note['main']:
                    name = rename_note(synd,note['Name'][:-2].split('","')[1],path,num)
                else:
                    name = note['Name'][:-2].split('","')[1]
            
                f_id,p_link = move_to(synd,f"{path}/{name}",dest)
                if not note['Original'] in ['',None]:
                    org_name = rename_note(synd,note['Original'],path,num)
                    move_to(synd,f"{path}/{org_name}",dest)
     
            if data['create_folder']: p_link = path_link

            src = re.findall('\d+',num)
            only_num = int(src[0]) if src else num
            if data['create_folder']:
                data['link'] = f'=HYPERLINK("#dlink=/d/f/{p_link}", "{only_num}")'
            else:
                data['link'] = f'=HYPERLINK("#dlink=/oo/r/{p_link}", "{only_num}")'

def dept_notes(PASS,reg_notes):
    config = txt2dict("config.txt")
    path = f"{config['despacho']}/Outbox Despacho"
    deps = {d.lower():d for d in config['deps'].split(",")}
    
    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        for num,data in reg_notes.items():
            if data['Dept'] in ['',None]:
                continue

            dest = "/team-folders/Mail {0}/Inbox {0}"
                
            dests = []
            for dep in data['Dept'].split(","):
                dp = dep.strip().lower() if dep != '' else ''
                if dp != '' and dp in deps.keys():
                    dests.append(dest.format(deps[dp]))
                
            for note in data['notes']:
                if note['main']:
                    name = rename_note(synd,note['Name'][:-2].split('","')[1],path,num)
                else:
                    name = note['Name'][:-2].split('","')[1]
                
                for dest in dests:
                    des,p_link = create_folder(synd,data,dest,num)
                    if dest == dests[-1]:
                        move_to(synd,f"{path}/{name}",des)
                    else:
                        copy_to(synd,f"{path}/{name}",des)


def upload_register(PASS,wb):
    config = txt2dict("config.txt")
    date = datetime.today().strftime('%d-%m-%Y-%HH-%Mm')
    
    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        file = NamedTemporaryFile()
        wb.save(file)
        file.seek(0)
        file.name = f"register-{date}.xlsx"
    
        print("Creating register file")
        uploaded = True
        
        try:
            ret_upload = synd.upload_file(file, dest_folder_path=f"{config['despacho']}/Outbox Despacho")
        except:
            print("Cannot upload register")
            wb.save(f"register-{date}.xlsx")
            uploaded = False

        if uploaded:
            try:
                ret_convert = synd.convert_to_online_office(ret_upload['data']['display_path'],
                    delete_original_file=False,
                    conflict_action='autorename')
            except:
                print("Cannot convert file to Synology Office")

        return True
    
    print("Cannot upload register")
    wb.save(f"register-{date}.xlsx")


def create_register(ws,reg_notes):
    reg_titles = ['source','link','Year','Ref','Date','Content','Dept','of_anex']
    ws.append(reg_titles)

    for num,data in dict(sorted(reg_notes.items())).items():
        if data['Dept'] in ['',None]:
            continue

        row = []
        for title in reg_titles[:-1]:
            row.append(data[title])

        n = len(data['notes']) - 1

        row.append(n) if n > 0  else row.append('')

        ws.append(row)
        ws[ws.max_row][4].value = datetime.strptime(data['Date'],"%d/%m/%Y")
        ws[ws.max_row][4].number_format = 'dd/mm/yyyy;@'
        
        column_widths = [10,10,10,12,12,50,12,12]
        for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
            ws.column_dimensions[get_column_letter(i)].width = column_width
            
        for row in ws[1:ws.max_row]:  # skip the header
            for i,col in enumerate(column_widths):
                cell = row[i]             # column H
                cell.alignment = Alignment(horizontal='center')

def fill_data(reg_notes):
    not_to_reg = ['Name','Original']
    for num,data in reg_notes.items():
        create_folder = False
        cont = 0
        for note in data['notes']:
            cont += 1
            for title in TITLES:
                if not title in data: data[title] = ''
                if not title in not_to_reg and note[title] != '':
                    data[title] = note[title]

        if cont > 1:
            create_folder = True
        else:
            nm = note["Name"].split(',')[1].split('"')[1]
            ext = Path(nm).suffix[1:]
            if not ext in EXT.values():
                create_folder = True

        data['create_folder'] = create_folder
                
            


def main():
    PASS = getpass()
    
    reg_notes = read_register(PASS)

    if reg_notes != {}:
        fill_data(reg_notes)
        
        #input("asd")

        archive_notes(PASS,reg_notes)
        dept_notes(PASS,reg_notes)
        
        wb = Workbook()
        ws = wb.active
        create_register(ws,reg_notes)
        upload_register(PASS,wb)

    
    input("Pulse Enter to continue")



if __name__ == '__main__':
    main()
