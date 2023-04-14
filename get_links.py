#!/bin/python

from getpass import getpass
from datetime import datetime
from tempfile import NamedTemporaryFile
import time
import re
from pathlib import Path

import pickle

from openpyxl import Workbook, load_workbook
from synology_drive_api.drive import SynologyDrive

from syno_tools import txt2dict, copy_to, convert_to


def get_ctr(folder):
    if folder[:7] == 'Mailbox':
        return folder[8:]
    elif folder[:10] == '8.Mailbox-':
        return folder[10:]


ext = {'xls':'osheet','xlsx':'osheet','docx':'odoc'}

def tests(PASS):
    config = txt2dict("config.txt")

    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        reg_file = synd.download_file('/mydrive/00 - Admin/reg.osheet')
        wb = load_workbook(reg_file)


        for ws_name in wb.sheetnames:
            ws = wb[ws_name]

            for row in ws.iter_rows(values_only=True):
                num = row[0]

                notes = synd.list_folder('/team-folders/Archive/cg in 2023')['data']['items']

                print(notes)


def main():
    PASS = getpass()
   
    tests(PASS)


if __name__ == '__main__':
    main()
