#!/bin/python

from getpass import getpass
from datetime import datetime
from tempfile import NamedTemporaryFile
import time
import re
from pathlib import Path
import ast

import pickle

from openpyxl import Workbook
from synology_drive_api.drive import SynologyDrive

from syno_tools import txt2dict, copy_to, move_to, convert_to, build_link, EXT


def get_notes_in_folders(PASS):
    config = txt2dict("config.txt")

    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd: 
        team_folders = synd.get_teamfolder_info()

        reg_notes = {}
        for folder in team_folders:
            if folder[:5] == 'Mail ':
                key = folder[5:]
            
                # Cheking all notes ########################################
                if not key in ['asr','cg','r','ctr','vc']:
                    print(f"Checking folder {folder}")
                    try:
                        notes = synd.list_folder(f"/team-folders/{folder}/Outbox {key}")['data']['items']
                        for note in notes:
                            print(f"    ->Found note {note['name']}")
                            reg_notes[note['name']] = note|{'source':key,'converted':False,'original':''}
                    except:
                        print(f'Cannot get notes from {folder}')
                        continue

        return reg_notes

def create_register(ws,reg_notes):
    year = datetime.today().strftime('%Y')
    ws.append(['type','dr','No','Year','Ref','Date','Content','Dept','link','Original'])
    
    for name,note in reg_notes.items():
        num = re.findall('\d+',name)
        num = f"000{num[0]}"[-4:] if num else ''
        if num and note['type'] == 'ctr in': num = num[1:]
        note['num'] = num

        note['year'] = year

        if note['type'] == 'r in':
            src = re.findall('\D+',name)
            note['source'] = src[0] if src else ''
        
        if note['converted']: continue
        
        nm = note['link'] if 'link' in note else name

        ws.append([note['type'],note['source'],num,year,'','','','',nm,note['original']])    
        ws[ws.max_row][5].value = datetime.today()
        ws[ws.max_row][5].number_format = 'dd/mm/yyyy;@'


def change_names(PASS,notes):
    config = txt2dict("config.txt")

    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        new_names = []
        for name,note in notes.items():
            try:
                #new_name = name if note['source'] in name else f"{note['source']}_{name}"
                new_name = name.strip()
                #new_name = new_name.replace(' ','_')
                if name != new_name:
                    new_names.append([new_name,name])
                    synd.rename_path(new_name,f"{note['display_path']}")
                    print(f"Name change from {name} to {new_name}") 
            except:
                print(f"ERROR: Cannot change name of {name}")

        for new in new_names:
            notes[new[0]] = notes.pop(new[1])

def move_to_ToSend(PASS,notes):
    config = txt2dict("config.txt")
    print("Moving")

    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        for name,note in notes.items():
            print(f"Note {name}")
            try:
                name_link = name
                f_id,p_link = move_to(synd,f"{note['display_path']}",f"/mydrive/ToSend")
                note['folder'] = f"/mydrive/ToSend"
                ext = Path(name).suffix[1:]
                if ext in EXT.values():
                    note['link'] = build_link(p_link,name_link)
                else:
                    note['link'] = build_link(p_link,name_link,True)
            except:
                print(f"Cannot move {name}")


def convert_files(PASS,notes):
    config = txt2dict("config.txt")

    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        new_names = []
        for name,note in notes.items():
            try:
                name_link = name
                note['converted'] = False
                # Here I check if I can convert the file to synology with the extension
                ext = Path(name).suffix[1:]
                if ext in EXT:
                    f_path,f_id,p_link = convert_to(synd,f"/mydrive/ToSend/{note['name']}")
                    #note['converted'] = True
                    note['original'] = name
                    note['link'] = build_link(p_link,name_link)
                    new_names.append([f"{name[:-len(ext)]}{EXT[ext]}",name])
            except:
                print(f"Cannot convert {name}")

        for new in new_names:
            notes[new[0]] = notes.pop(new[1])



def upload_register(PASS,wb):
    config = txt2dict("config.txt")
    date = datetime.today().strftime('%d-%m-%Y-%HH-%mm')
    
    with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        file = NamedTemporaryFile()
        wb.save(file)
        file.seek(0)
        file.name = f"{date}-ctr-incoming.xlsx"
    
        print("Creating register file")
        uploaded = True
        
        try:
            ret_upload = synd.upload_file(file, dest_folder_path=f"/mydrive/ToSend")
        except:
            print("Cannot upload register")
            wb.save(f"from_dr-{date}.xlsx")
            uploaded = False

        if uploaded:
            try:
                ret_convert = synd.convert_to_online_office(ret_upload['data']['display_path'],
                    delete_original_file=True,
                    conflict_action='autorename')
            except:
                print("Cannot convert file to Synology Office")

        return True
    
    print("Cannot upload register")
    wb.save(f"from_dr-{date}.xlsx")




def main():
    PASS = getpass()
    
    reg_notes = get_notes_in_folders(PASS)
    input("go") 
    if reg_notes != {}:
        wb = Workbook()
        ws = wb.active
        
        try:
            change_names(PASS,reg_notes)
            move_to_ToSend(PASS,reg_notes)
            convert_files(PASS,reg_notes)
            create_register(ws,reg_notes)
            upload_register(PASS,wb)
        except:
            create_register(ws,reg_notes)
            upload_register(PASS,wb)
    
    input("Pulse Enter to continue")



if __name__ == '__main__':
    main()
