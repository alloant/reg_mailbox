#!/bin/python

from getpass import getpass
from datetime import datetime
from tempfile import NamedTemporaryFile
import time
import re

from pathlib import Path

import logging

import ast
import pickle

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from synology_drive_api.drive import SynologyDrive

from synomail import CONFIG, EXT
from synomail.syno_tools import move_to, convert_to


def folder_in_teams(folder,teams):
    fds = folder.split("/")[1:]
    for team in teams:
        tms = team['folder'].split("/")[1:]
        same = True
        key = team['type']
        for i,fd in enumerate(fds):
            if '@' in tms[i]:
                pt = tms[i].replace('@','')
                if fds[i][:len(pt)] == pt:
                    key = fds[i][len(pt):]
                else:
                    same = False
                    break
            else:
                if fds[i] != tms[i]:
                    same = False
                    break

        if same:
            nt = team.copy()
            nt['folder'] = team['folder'].replace('@',key)
            return same,key,nt

    return False,'asd',''


def get_notes_in_folders(PASS):
    year = datetime.today().strftime('%Y')

    with SynologyDrive(CONFIG['user'],PASS,"nas.prome.sg",dsm_version='7') as synd: 
        team_folders = synd.get_teamfolder_info()
        team_config = [ast.literal_eval(tm) for tm in CONFIG['teams'].split('|')]

        reg_notes = {}
        for folder in team_folders:
            mail_folder,key,team = folder_in_teams(f"/team-folders/{folder}",team_config)
            
            if mail_folder: # and key == 'gul':
                logging.debug(f"Checking folder {team['folder']}")
                try:
                    notes = synd.list_folder(team['folder'])['data']['items']
                    for note in notes:
                        logging.info(f"Found note {note['name']} in {team['folder']}")
                        reg_notes[note['name']] = team.copy()|{'source':key,'converted':False,'original':'','p_link':note['permanent_link']}
                except Exception as err:
                    logging.error(err)
                    logging.error(f'Cannot get notes from {team["folder"]}')
                    continue

        return reg_notes

def create_register(ws,reg_notes):
    year = datetime.today().strftime('%Y')
    ws.append(['type','source','No','Year','Ref','Date','Content','Dept','Name','Original','Comments'])
    
    #for name,note in dict(sorted(reg_notes.items())).items():
    #if 'No' in reg_notes:
    #    ord_notes = dict(sorted(reg_notes.items(), key=lambda item: item[1]['No']))
    #if 'source' in reg_notes:
    #    ord_notes = dict(sorted(ord_notes.items(), key=lambda item: item[1]['source']))
    #if 'type' in reg_notes:
    #    ord_notes = dict(sorted(ord_notes.items(), key=lambda item: item[1]['type']))

    for name,note in reg_notes.items():
        num = re.findall('\d+',name.replace(note['source'],''))
        num = num[0] if num else ''
        
        #num = f"000{num[0]}"[-4:] if num else ''
        #if num and note['type'] == 'ctr in': num = num[1:]
        
        note['num'] = num

        note['year'] = year

        if note['type'] == 'r in':
            src = re.findall('\D+',name)
            note['source'] = src[0] if src else ''
        
        if note['converted']: continue
        
        nm = note['link'] if 'link' in note else name

        ws.append([note['type'],note['source'],num,year,'','','','',nm,note['original'],''])    
        ws[ws.max_row][5].value = datetime.today()
        ws[ws.max_row][5].number_format = 'dd/mm/yyyy;@'

        column_widths = [10,10,10,10,12,12,50,12,20,20]
        for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
            ws.column_dimensions[get_column_letter(i)].width = column_width

        for row in ws[1:ws.max_row]:  # skip the header
            for i,col in enumerate(column_widths):
                cell = row[i]             # column H
                cell.alignment = Alignment(horizontal='center')

def change_names(PASS,notes):
    with SynologyDrive(CONFIG['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        new_names = []
        for name,note in notes.items():
            try:
                new_name = name if note['source'] in name or note['source'] in ['r','cg'] else f"{note['source']}_{name}"
                new_name = new_name.strip()
                #new_name = new_name.replace(' ','_')
                new_names.append([new_name,name])
                synd.rename_path(new_name,f"{note['folder']}/{name}")
            except Exception as err:
                logging.error(err)
                logging.warning(f"Cannot change name of {name}")

        for new in new_names:
            notes[new[0]] = notes.pop(new[1])


def move_to_despacho(PASS,notes):
    with SynologyDrive(CONFIG['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        for name,note in notes.items():
            try:
                ext = Path(name).suffix[1:]
                name_link = name
                
                move_to(synd,f"{note['folder']}/{name}",f"{CONFIG['despacho']}/Inbox Despacho")
                
                note['folder'] = f"{CONFIG['despacho']}/Inbox Despacho"

                chain = 'oo/r' if ext in EXT.values() else 'd/f'
                
                p_link = note['p_link']
                link = f'=HYPERLINK("#dlink=/{chain}/{p_link}", "{name_link}")' if p_link != '' else name_link
                note['link'] = link
            except Exception as err:
                logging.error(err)


def convert_files(PASS,notes):
    with SynologyDrive(CONFIG['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        new_names = []
        for name,note in notes.items():
            try:
                name_link = name
                note['converted'] = False
                # Here I check if I can convert the file to synology with the extension
                ext = Path(name).suffix[1:]
                if ext in EXT:
                    f_path,f_id,p_link = convert_to(synd,f"{CONFIG['despacho']}/Inbox Despacho/{name}")
                    
                    note['original'] = name
                    new_name = f"{name[:-len(ext)]}{EXT[ext]}"
                    new_names.append([new_name,name])
                    if p_link != '':
                        note['link'] = f'=HYPERLINK("#dlink=/oo/r/{p_link}", "{new_name}")'

            except Exception as err:
                logging.error(err)

        for new in new_names:
            notes[new[0]] = notes.pop(new[1])



def upload_register(PASS,wb):
    date = datetime.today().strftime('%Y-%m-%d')
    hour = datetime.today().strftime('%HH-%Mm')
    
    with SynologyDrive(CONFIG['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        file = NamedTemporaryFile()
        wb.save(file)
        file.seek(0)
        file.name = f"despacho-{date}-{hour}.xlsx"
    
        logging.info("Creating register file")
        uploaded = True
        
        try:
            ret_upload = synd.upload_file(file, dest_folder_path=f"{CONFIG['despacho']}/Inbox Despacho")
        except Exception as err:
            logging.error(err)
            logging.error("Cannot upload register")
            wb.save(f"despacho-{date}-{hour}.xlsx")
            uploaded = False

        if uploaded:
            try:
                ret_convert = synd.convert_to_online_office(ret_upload['data']['display_path'],
                    delete_original_file=False,
                    conflict_action='autorename')
            except Exception as err:
                logging.error(err)
                logging.warning("Cannot convert register to Synology Office")

            return True
    
    logging.error("Cannot upload register")
    wb.save(f"despacho-{date}-{hour}.xlsx")

def init_get_mail(PASS):
    logging.info('Starting searching new mail')
    reg_notes = get_notes_in_folders(PASS)
    
    if reg_notes != {}:
        wb = Workbook()
        ws = wb.active
        
        try:
            change_names(PASS,reg_notes)
            move_to_despacho(PASS,reg_notes)
            convert_files(PASS,reg_notes)
            create_register(ws,reg_notes)
            upload_register(PASS,wb)
        except:
            create_register(ws,reg_notes)
            upload_register(PASS,wb)

    logging.info('Finish searching new mail')
    
    

def main():
    PASS = getpass()
    init_get_mail(PASS)
    input("Pulse Enter to continue")


if __name__ == '__main__':
    main()
