#!/bin/python

from getpass import getpass
from datetime import datetime
from tempfile import NamedTemporaryFile
import time
import re
from pathlib import Path
import ast

import logging

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

from synochat.webhooks import IncomingWebhook
from synology_drive_api.drive import SynologyDrive

from synomail.syno_tools import move_to, convert_to, upload_convert_wb
from synomail.get_mail import create_despacho
from synomail import CONFIG, EXT

TITLES = ['type','source','No','Year','Ref','Date','Content','Dept','Name','Original','Comments'] 

def read_register(PASS):
    with SynologyDrive(CONFIG['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        register = synd.list_folder(f"{CONFIG['despacho']}/Outbox Despacho")['data']['items']
        pattern = re.compile("despacho-(\d\d-\d\d-\d\d\d\d)-(\d\dH-\d\dm).osheet")
        notes = {}
        
        for reg in register:
            if reg["name"][:9] == "despacho-" and 'osheet' in reg['name']:
                try:
                    reg_file = synd.download_file(reg['display_path'])
                    wb = load_workbook(reg_file)
                    
                    for ws_name in wb.sheetnames:
                        ws = wb[ws_name]
                        
                        for row in ws.iter_rows(values_only=True):
                            if not row[2] in ['',None] and not row[1] in ['',None] and row[0] != 'type':
                                n = f"000{row[2]}"[-4:]
                                notes[f"{row[1]}_{n}"] = {'notes':[]}
                       
                        for row in ws.iter_rows(values_only=True):
                            if not row[2] in ['',None] and not row[1] in ['',None] and row[0] != 'type':
                                temp = {}
                                for i,title in enumerate(TITLES):
                                    if not title in temp: temp[title] = ''

                                    if not row[i] in ['',None]:
                                        temp[title] = row[i]

                                    if title == 'No':
                                        temp[title] = f"000{row[i]}"[-4:]
                                    if isinstance(row[i],datetime):
                                        temp[title] = row[i].strftime('%d/%m/%Y')

                                temp['main'] = True if temp['Content'] != '' else False

                                notes[f"{temp['source']}_{temp['No']}"]['notes'].append(temp.copy())
                except Exception as err:
                    logging.error(err)
                    logging.error(f"Cannot read register")
                    continue

        return notes


def create_folder(synd,data,dest,num):
    if data['create_folder']:
        try:
            new_folder = synd.get_file_or_folder_info(f"{dest}/{num}")
            exist = True
        except:
            exist = False

        try:
            if not exist:
                rst = synd.create_folder(num, dest)
                p_link = rst['data']['permanent_link']
            else:
                p_link = new_folder['data']['permanent_link']

            dest = f"{dest}/{num}"

            return dest,p_link
        except:
            logging.error(f"Cannot create folder {num}")
            return '',''
    else:
        return dest,''


def rename_note(synd,note_name,path,num):
        #name = note_name[:-2].split('","')[1]
        name = note_name
        try:
            ext = Path(name).suffix[1:]
            synd.rename_path(f"{num}.{ext}",f"{path}/{name}")
            name = f"{num}.{ext}"
        except:
            logging.error("Cannot change the name")
        
        return name

        
def archive_notes(PASS,reg_notes):
    path = f"{CONFIG['despacho']}/Outbox Despacho"
    with SynologyDrive(CONFIG['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        for num,data in reg_notes.items():
            if len(data['notes']) == 0:
                continue

            if data['Dept'] in [None,'']:
                data['link'] = ''
                continue

            dest = f"{CONFIG['archive']}/{data['type']} in {data['Year']}"
            dest,path_link = create_folder(synd,data,dest,num)
            
            for note in data['notes']:
                if note['main']:
                    name = rename_note(synd,note['Name'][:-2].split('","')[1],path,num)
                else:
                    name = note['Name'][:-2].split('","')[1]

            
                move_to(synd,f"{path}/{name}",dest)
                ext = Path(f"{path}/{name}").suffix[1:]
                
                if not note['Original'] in ['',None]:
                    org_name = rename_note(synd,note['Original'],path,num)
                    move_to(synd,f"{path}/{org_name}",dest)

     
            src = re.findall('\d+',num)
            only_num = int(src[0]) if src else num
            
            if data['create_folder']:
                data['link'] = f'=HYPERLINK("#dlink=/d/f/{path_link}", "{only_num}")'
                data['link_dep'] = f'<https://nas.prome.sg:5001/d/f/{path_link}|{name}>'
                data['link_dep_2'] = f'<https://nas.prome.sg:5001/d/f/{path_link}|{data["source"]} {only_num}/{data["Year"][2:]}>'
            elif ext in EXT:
                p_link = note['Name'][:-2].split('","')[0].split("/")[-1]
                data['link'] = f'=HYPERLINK("#dlink=/oo/r/{p_link}", "{only_num}")'
                data['link_dep'] = f'<https://nas.prome.sg:5001/oo/r/{p_link}|{name}>'
                data['link_dep_2'] = f'<https://nas.prome.sg:5001/oo/r/{p_link}|{data["source"]} {only_num}/{data["Year"][2:]}>'

            else:
                p_link = note['Name'][:-2].split('","')[0].split("/")[-1]
                data['link'] = f'=HYPERLINK("#dlink=/d/f/{p_link}", "{only_num}")'
                data['link_dep'] = f'<https://nas.prome.sg:5001/d/f/{p_link}|{name}>'
                data['link_dep_2'] = f'<https://nas.prome.sg:5001/d/f/{p_link}|{data["source"]} {only_num}/{data["Year"][2:]}>'



def upload_register(PASS,wb,name,dest):        
    with SynologyDrive(CONFIG['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
        upload_convert_wb(synd,wb,name,dest) 
    
def create_register(ws,reg_notes):
    reg_titles = ['source','link','Year','Ref','Date','Content','Dept','of_anex']
    ws.append(reg_titles)

    font = Font(name= 'Arial',
                size=12,
                bold=False,
                italic=False,
                strike=False,
                underline='none'
                #color='4472C4'
                )

    #for num,data in dict(sorted(reg_notes.items())).items():
    #if 'No' in reg_notes:
    #    ord_notes = dict(sorted(reg_notes.items(), key=lambda item: item[1]['No']))
    #if 'source' in reg_notes:
    #    ord_notes = dict(sorted(ord_notes.items(), key=lambda item: item[1]['source']))
    #if 'type' in reg_notes:
    #    ord_notes = dict(sorted(ord_notes.items(), key=lambda item: item[1]['type']))

    for num,data in reg_notes.items():
        if len(data['notes']) == 0:
            continue

        if data['Dept'] in ['',None]:
            continue

        row = []
        for title in reg_titles[:-1]:
            row.append(data[title])

        n = len(data['notes']) - 1

        row.append(n) if n > 0  else row.append('')

        ws.append(row)
        ws[ws.max_row][4].value = datetime.strptime(data['Date'],"%d/%m/%Y")
        ws[ws.max_row][4].number_format = 'dd/mm/yyyy;@'
        
        column_widths = [10,10,10,12,12,50,12,12]
        for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
            ws.column_dimensions[get_column_letter(i)].width = column_width
            
        for row in ws[1:ws.max_row]:  # skip the header
            for i,col in enumerate(column_widths):
                cell = row[i]             # column H
                cell.alignment = Alignment(horizontal='center')
                cell.font = font

def fill_data(reg_notes):
    not_to_reg = ['Name','Original']
    for num,data in reg_notes.items():
        if len(data['notes']) == 0:
            continue

        create_folder = False
        cont = 0
        for note in data['notes']:
            cont += 1
            for title in TITLES:
                if not title in data: data[title] = ''
                if not title in not_to_reg and note[title] != '':
                    data[title] = note[title]

        if cont > 1:
            create_folder = True
        """
        else:
            nm = note["Name"].split(',')[1].split('"')[1]
            ext = Path(nm).suffix[1:]
            if not ext in EXT.values():
                create_folder = True
        """

        data['create_folder'] = create_folder
                

def send_messages(reg_notes):
    tokens = ast.literal_eval(CONFIG['tokens'])
    
    for num,data in reg_notes.items():
        if len(data['notes']) == 0:
            continue

        deps = [dp.lower() for dp in data['Dept'].split(',')]
        
        message = f"Assigned to: *{data['Dept']}* \nLink: {data['link_dep_2']} \nContent: `{data['Content']}`"
        
        if data['Ref'] != '':
            message += f"\nRef: _{data['Ref']}_"
        
        if data['Comments'] != '':
            message += f"\nComment: _{data['Comments']}_"
 
        message +=  f"\nRegistry date: {data['Date']}"

        for dep in deps:
            webhook = IncomingWebhook('nas.prome.sg', tokens[dep], port=5001)
            webhook.send(message)

def init_register_despacho(PASS):
    logging.info('Starting register despacho')
    reg_notes = read_register(PASS)

    date = datetime.today().strftime('%Y-%m-%d-%HH-%Mm')
    name = f"register-{date}.xlsx"

    if reg_notes != {}:
        fill_data(reg_notes)
        
        archive_notes(PASS,reg_notes)
        send_messages(reg_notes)
        #print(reg_notes) 
        wb_reg = Workbook()
        ws_reg = wb_reg.active
        create_register(ws_reg,reg_notes)
        upload_register(PASS,wb_reg,name,f"{CONFIG['despacho']}/Outbox Despacho")
        #print(reg_notes)
        #wb_des = Workbook()
        #ws_des = wb_des.active
        #create_despacho(ws_des,reg_notes)
        #upload_register(PASS,wb_des,name.replace('register','despacho'),f"{CONFIG['despacho']}/Outbox Despacho")

    logging.info('Finishing register despacho')

def main():
    PASS = getpass()
    init_register_despacho(PASS)
    
    input("Pulse Enter to continue")



if __name__ == '__main__':
    main()
