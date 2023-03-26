#!/bin/python

from getpass import getpass
from datetime import datetime
from tempfile import NamedTemporaryFile

from openpyxl import Workbook, load_workbook
from synology_drive_api.drive import SynologyDrive


def txt2dict(file):
    DICT = {}
    
    with open(file,mode='r') as inp:
        lines = inp.read().splitlines()

        DICT = {ln.split(':')[0]:ln.split(':')[1] for ln in lines if ":" in ln}
    
    return DICT

def get_ctr(folder):
    if folder[:7] == 'Mailbox':
        return t[8:]
    elif folder[:10] == '8.Mailbox-':
        return t[10:]

def get_link(filefolder,text):
    p_path = filefolder['data']['permanent_link']

    if filefolder['data']['type'] == 'file':
        return f'=HYPERLINK("#dlink=/oo/r/{p_path}", "{text}")'
    else:
        return f'=HYPERLINK("#dlink=/d/f/{p_path}", "{text}")'


def save_synology(synd,wb,name,dest="/mydrive"):
    file = NamedTemporaryFile()
    wb.save(file)
    file.seek(0)
    file.name = name
    
    print("Creating register file")
    uploaded = True
    try:
        ret_upload = synd.upload_file(file, dest_folder_path=dest)
    except:
        print("Cannot upload register")
        wb.save(name)
        uploaded = False

    if uploaded:
        try:
            ret_convert = synd.convert_to_online_office(ret_upload['data']['display_path'],
                delete_original_file=True,
                conflict_action='autorename')
        except:
            print("Cannot convert file to Synology Office")



config = txt2dict("config.txt")
groups = txt2dict("groups.txt")


year = datetime.today().strftime('%Y')
date = datetime.today().strftime('%d/%m/%Y')

MAIL = {}
REPORT = []

with open("ctrs.txt",mode='r') as inp:
    MAIL = {ln:[] for ln in inp.read().splitlines() if ln != ''}

    output_wb = load_workbook(filename = "output.xlsx")
    output = output_wb.active
    
    cols = []
    for row in output.iter_rows():
        for ctr in str(row[0].value).split(','):
            nt = {}
            header = True
            for col in range(1,len(row)):
                if ctr == 'ctr':
                    cols.append(row[col].value)
                else:
                    nt[cols[col-1]] = row[col].value
                    header = False
            
            if not header and ctr != '':
                if ctr in groups:
                    for c in groups[ctr].split(','):
                        MAIL[c].append(nt)
                else:
                    if ctr in MAIL:
                        MAIL[ctr].append(nt)
                    else:
                        print(f"ctr doesn't exist")
        
        if row[0].value != 'ctr': REPORT.append([row[0].value]+list(nt.values()))


PASS = getpass()
with SynologyDrive(config['user'],PASS,"nas.prome.sg",dsm_version='7') as synd:
    #mail = synd.list_folder(f"{config['archive']}/ctr out {year}")
    for ctr,notes in MAIL.items():
        if len(notes) == 0: continue

        wb = Workbook()
        ws = wb.active

        ws.append(['ctr','No','Year','Ref','Date','Content'])
        for note in notes:
            link = get_link(synd.get_file_or_folder_info(f"{config['archive']}/ctr out {year}/{note['No']}"),note['No'])

            ws.append([ctr,link,year,'',date,note['Content']])
        
        try:
            save_synology(synd,wb,f"{date.replace('/','-')}-mail-from-cg.xlsx",f"/team-folders/Mailbox {ctr}/cr to {ctr}")
        except:
            save_synology(synd,wb,f"{date.replace('/','-')}-mail-from-cg.xlsx",f"/team-folders/8.Mailbox-{ctr}/cr-to-{ctr}")

    wb = Workbook()
    ws = wb.active
    ws.append(['ctr','No','Year','Ref','Date','Content'])
    for r in REPORT:
        ws.append(r)

    try:
        save_synology(synd,wb,f"{date.replace('/','-')}-out-ctr.xlsx",f"/mydrive")

